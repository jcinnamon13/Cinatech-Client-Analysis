"""
CinaTech Client Onboarding Template Generator
==============================================
Rebuilds public/CinaTech_Client_Onboarding_Template.xlsx with:
- 9 analysis pillars (matching lib/ai/prompts.ts)
- 2–3 targeted questions per pillar
- CinaTech branded title row
- Colour-coded pillar headers
- Light blue input cells with thin borders
- Italic guidance text per question
- Completeness indicator formula with conditional formatting
- Data validation dropdowns for delivery model and sector

Prerequisites:
    pip3 install --user openpyxl

Usage:
    python3 scripts/build_onboarding_template.py

Output:
    public/CinaTech_Client_Onboarding_Template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------------------------------------------------------------------------
# Output path
# ---------------------------------------------------------------------------
OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', 'public', 'CinaTech_Client_Onboarding_Template.xlsx'
)

# ---------------------------------------------------------------------------
# Colour constants (openpyxl uses hex without '#')
# ---------------------------------------------------------------------------
C_TITLE_BG       = "1F2D45"   # dark navy
C_PILLAR_BG      = "2E4A7A"   # mid navy
C_QUESTION_BG    = "F0F4FA"   # very light blue-grey
C_INPUT_BG       = "D6E4F0"   # light blue
C_GUIDANCE_BG    = "F9F9F9"   # near-white
C_WHITE          = "FFFFFF"
C_DARK           = "1A1A2E"
C_GREY           = "888888"
C_SPACER         = "EEEEEE"
C_GREEN_BG       = "D4EDDA"
C_GREEN_FG       = "155724"
C_AMBER_BG       = "FFF3CD"
C_AMBER_FG       = "856404"
C_RED_BG         = "F8D7DA"
C_RED_FG         = "721C24"

# ---------------------------------------------------------------------------
# Pillar data
# ---------------------------------------------------------------------------
PILLARS = [
    {
        "number": 1,
        "title": "Competitive Positioning and Differentiation",
        "questions": [
            {
                "label": "P1 Q1 — Competitors",
                "prompt": "Who are your top 3 competitors and how do clients typically compare you to them?",
                "guidance": "Name them directly. Note how prospects frame the comparison — price, quality, specialisation, speed, or reputation.",
                "placeholder": "[List competitors and how prospects compare you to them]",
                "validation": None,
            },
            {
                "label": "P1 Q2 — Differentiation",
                "prompt": "What is your primary point of differentiation? (price, quality, specialisation, speed, or results guarantee)",
                "guidance": "Choose one main differentiator and explain it briefly. What would a client say if asked why they chose you over a competitor?",
                "placeholder": "[Describe your primary differentiator and why clients choose you]",
                "validation": None,
            },
            {
                "label": "P1 Q3 — Competitor gaps",
                "prompt": "What do competitors offer that you currently do not?",
                "guidance": "Be honest. This is not a weakness to hide — it is intelligence the agency needs to give you accurate strategic advice.",
                "placeholder": "[List gaps honestly — services, reach, pricing, technology, etc.]",
                "validation": None,
            },
        ],
    },
    {
        "number": 2,
        "title": "Target Market Definition and Ideal Client Profile",
        "questions": [
            {
                "label": "P2 Q1 — Ideal client",
                "prompt": "Describe your ideal client — industry, company size, geography, and decision-maker role.",
                "guidance": "Be specific. 'SMEs in the UK' is not enough. Example: 'Owner-operated aesthetics clinics in England, 2–10 staff, decision-maker is the clinic owner directly.'",
                "placeholder": "[Industry / company size / geography / decision-maker role]",
                "validation": None,
            },
            {
                "label": "P2 Q2 — Core problem",
                "prompt": "What is the single biggest problem your best clients had before working with you?",
                "guidance": "State it in the client's own words if possible. This feeds directly into the Jobs-to-be-Done analysis.",
                "placeholder": "[The core problem your best clients had — in their own words if possible]",
                "validation": None,
            },
            {
                "label": "P2 Q3 — Profitable segments",
                "prompt": "Which client segments are most profitable for you, and which do you want more of?",
                "guidance": "These may not be the same. Distinguish between your current best clients and your target best clients.",
                "placeholder": "[Most profitable segment now / segment you want more of]",
                "validation": None,
            },
        ],
    },
    {
        "number": 3,
        "title": "Client Acquisition and Growth Engine",
        "questions": [
            {
                "label": "P3 Q1 — Lead sources",
                "prompt": "What are your top 3 lead sources by volume and by revenue?",
                "guidance": "Volume and revenue rankings often differ — a referral brings fewer leads but higher value. List both rankings separately.",
                "placeholder": "[Top 3 by volume / Top 3 by revenue]",
                "validation": None,
            },
            {
                "label": "P3 Q2 — CAC and client value",
                "prompt": "What is your average cost to acquire a new client and your average first-year client value (£)?",
                "guidance": "Rough estimates are fine. Include time cost if there is no fixed ad spend. CAC and LTV are foundational to the Ansoff Matrix analysis.",
                "placeholder": "[Average CAC £ / Average first-year client value £]",
                "validation": None,
            },
            {
                "label": "P3 Q3 — Conversion rate",
                "prompt": "What is your current lead-to-client conversion rate, and where do most prospects drop off?",
                "guidance": "Identify the drop-off stage: awareness, enquiry, proposal, or close. Even a rough percentage is useful.",
                "placeholder": "[Conversion rate % / Stage where prospects most commonly drop off]",
                "validation": None,
            },
        ],
    },
    {
        "number": 4,
        "title": "Onboarding Friction and Operational Readiness",
        "questions": [
            {
                "label": "P4 Q1 — Onboarding process",
                "prompt": "Walk us through your current client onboarding process. Where are the most common delays or friction points?",
                "guidance": "List each step in sequence. Mark any step that regularly causes delays, confusion, or client complaints.",
                "placeholder": "[Step-by-step onboarding process / friction points or common delays]",
                "validation": None,
            },
            {
                "label": "P4 Q2 — Tools and manual steps",
                "prompt": "What tools and systems underpin your service delivery? List any manual steps you would like to automate.",
                "guidance": "Include CRM, project management, communication, billing, and reporting tools. Note any spreadsheet-based or email-based processes.",
                "placeholder": "[Tools and systems used / manual steps you want to automate]",
                "validation": None,
            },
        ],
    },
    {
        "number": 5,
        "title": "Proof of Results and Case Studies",
        "questions": [
            {
                "label": "P5 Q1 — Client outcomes",
                "prompt": "Do you have documented client outcomes, case studies, or testimonials? Describe 1–2 specific results with numbers.",
                "guidance": "Include percentages, revenue figures, time saved, or leads generated. Vague praise ('great results') is far less useful than specific numbers.",
                "placeholder": "[Outcome 1 with numbers / Outcome 2 with numbers]",
                "validation": None,
            },
            {
                "label": "P5 Q2 — Proof in sales",
                "prompt": "How do you currently present proof of results to prospects during the sales process?",
                "guidance": "Examples: case study PDF, video testimonials, live results dashboard, written reference, reference call. Be specific about the format and timing.",
                "placeholder": "[How proof is presented to prospects — format and timing]",
                "validation": None,
            },
        ],
    },
    {
        "number": 6,
        "title": "Service Delivery Model and Scalability",
        "questions": [
            {
                "label": "P6 Q1 — Delivery model",
                "prompt": "How is your service delivered? Select from the dropdown.",
                "guidance": "Done-for-you = you do the work. Done-with-you = collaborative. Self-serve = client uses your platform/tools. Hybrid/Mixed = combination.",
                "placeholder": "[Select delivery model from dropdown]",
                "validation": "delivery_model",
            },
            {
                "label": "P6 Q2 — Maximum capacity",
                "prompt": "What is your maximum client capacity at current headcount, without further hiring?",
                "guidance": "Estimate based on hours per client per week × available team hours. Include yourself in the calculation if you are the primary delivery resource.",
                "placeholder": "[Maximum number of clients at current headcount]",
                "validation": None,
            },
            {
                "label": "P6 Q3 — Scalability bottleneck",
                "prompt": "What is the single biggest bottleneck to scaling your delivery?",
                "guidance": "Be specific: is it time, a particular skill, a missing system, or capital? Identifying one root constraint is more valuable than listing several.",
                "placeholder": "[Single biggest delivery bottleneck and why]",
                "validation": None,
            },
        ],
    },
    {
        "number": 7,
        "title": "Revenue Model and Growth Trajectory",
        "questions": [
            {
                "label": "P7 Q1 — Revenue streams",
                "prompt": "List your current revenue streams and the approximate percentage contribution of each.",
                "guidance": "Example: Retainer 60%, one-off projects 30%, training/courses 10%. If you only have one stream, note that clearly.",
                "placeholder": "[Revenue stream / % contribution — e.g. Retainer 60%, Projects 30%, Training 10%]",
                "validation": None,
            },
            {
                "label": "P7 Q2 — Revenue figures",
                "prompt": "What was your total revenue in the last 12 months, and what is your target for the next 12?",
                "guidance": "Exact or approximate figures in £. If revenue varies significantly month-to-month, note the range.",
                "placeholder": "[Last 12 months revenue £ / Target next 12 months £]",
                "validation": None,
            },
            {
                "label": "P7 Q3 — Recurring contracts",
                "prompt": "Do you offer recurring or retainer contracts? What is your average contract length and monthly churn rate?",
                "guidance": "If you do not track churn formally, estimate: how many clients cancel per quarter out of your total active client base?",
                "placeholder": "[Contract type / average length in months / monthly churn % or estimate]",
                "validation": None,
            },
        ],
    },
    {
        "number": 8,
        "title": "Client Education and Retention",
        "questions": [
            {
                "label": "P8 Q1 — Post-sign education",
                "prompt": "What structured onboarding, training, or education do you provide to clients after they sign?",
                "guidance": "Include kick-off calls, welcome sequences, resource libraries, portal access, check-in cadence. 'Nothing formal' is a valid answer.",
                "placeholder": "[Structured education / onboarding steps provided after sign-up]",
                "validation": None,
            },
            {
                "label": "P8 Q2 — At-risk management",
                "prompt": "How do you identify and manage at-risk clients? What is your average client lifespan in months?",
                "guidance": "Name your early-warning signals: missed calls, low engagement, payment delays, complaints. Describe what action you take when you spot them.",
                "placeholder": "[At-risk signals / management process / average client lifespan in months]",
                "validation": None,
            },
        ],
    },
    {
        "number": 9,
        "title": "Regulatory and Compliance Exposure",
        "questions": [
            {
                "label": "P9 Q1 — Sector",
                "prompt": "What sector do you operate in? Select from the dropdown.",
                "guidance": "Your sector determines which regulatory frameworks apply (CQC, MHRA, FCA, SRA, Ofsted, etc.). This is mandatory for compliance analysis.",
                "placeholder": "[Select your sector from the dropdown]",
                "validation": "sector",
            },
            {
                "label": "P9 Q2 — Guarantees",
                "prompt": "Do you make guarantees or promises of specific results in your marketing or contracts? Describe them exactly.",
                "guidance": "Copy the exact wording from your sales page, ads, or contract. Vague answers make compliance review impossible. 'We guarantee results' is not enough.",
                "placeholder": "[Exact wording of any guarantees or specific result promises, or 'None']",
                "validation": None,
            },
        ],
    },
]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def font_title():
    return Font(name="Calibri", size=14, bold=True, color=C_WHITE)

def font_pillar():
    return Font(name="Calibri", size=12, bold=True, color=C_WHITE)

def font_completeness_label():
    return Font(name="Calibri", size=10, bold=True, color=C_DARK)

def font_question():
    return Font(name="Calibri", size=11, bold=True, color=C_DARK)

def font_input():
    return Font(name="Calibri", size=11, color=C_DARK)

def font_guidance():
    return Font(name="Calibri", size=10, italic=True, color=C_GREY)

def align_center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def align_wrap(indent=0):
    return Alignment(horizontal="left", vertical="top", wrap_text=True, indent=indent)

def align_middle():
    return Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

# ---------------------------------------------------------------------------
# Build completeness formula
# ---------------------------------------------------------------------------

def build_completeness_formula(input_cells):
    """
    Builds a portable Excel formula that counts filled input cells.
    A cell is considered filled if it is non-empty AND does not start with '['.
    Uses an explicit IF chain for maximum compatibility across Excel versions.
    """
    parts = [
        f'IF({c}<>"",IF(LEFT({c},1)<>"[",1,0),0)'
        for c in input_cells
    ]
    return "=" + "+".join(parts)

# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Onboarding"

    # Column widths
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 56
    ws.column_dimensions['C'].width = 46

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # Sheet tab colour
    ws.sheet_properties.tabColor = C_TITLE_BG

    # -------------------------------------------------------------------
    # Row 1 — Title banner
    # -------------------------------------------------------------------
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "CinaTech  —  Strategic Client Onboarding Questionnaire"
    c.fill = fill(C_TITLE_BG)
    c.font = font_title()
    c.alignment = align_center()

    # -------------------------------------------------------------------
    # Row 2 — Completeness label
    # -------------------------------------------------------------------
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "SUBMISSION COMPLETENESS"
    c.fill = fill(C_QUESTION_BG)
    c.font = font_completeness_label()
    c.alignment = align_middle()
    c = ws["C2"]
    c.value = "Complete all 23 questions before submitting. The score below updates automatically."
    c.fill = fill(C_GUIDANCE_BG)
    c.font = font_guidance()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # -------------------------------------------------------------------
    # Row 3 — Completeness score (formula patched after loop)
    # -------------------------------------------------------------------
    ws.row_dimensions[3].height = 26
    ws.merge_cells("A3:B3")
    c = ws["A3"]
    c.value = 0  # placeholder — replaced below
    c.number_format = '"Responses Completed: "0" / 23"'
    c.font = Font(name="Calibri", size=12, bold=True, color=C_DARK)
    c.alignment = align_center()
    # Conditional formatting rules (highest priority first)
    ws.conditional_formatting.add(
        "A3:B3",
        FormulaRule(
            formula=["$A$3>=23"],
            fill=fill(C_GREEN_BG),
            font=Font(name="Calibri", size=12, bold=True, color=C_GREEN_FG),
        ),
    )
    ws.conditional_formatting.add(
        "A3:B3",
        FormulaRule(
            formula=["$A$3>=18"],
            fill=fill(C_AMBER_BG),
            font=Font(name="Calibri", size=12, bold=True, color=C_AMBER_FG),
        ),
    )
    ws.conditional_formatting.add(
        "A3:B3",
        FormulaRule(
            formula=["$A$3<18"],
            fill=fill(C_RED_BG),
            font=Font(name="Calibri", size=12, bold=True, color=C_RED_FG),
        ),
    )
    c = ws["C3"]
    c.fill = fill(C_GUIDANCE_BG)

    # -------------------------------------------------------------------
    # Row 4 — Spacer
    # -------------------------------------------------------------------
    ws.row_dimensions[4].height = 8
    for col in ["A", "B", "C"]:
        ws[f"{col}4"].fill = fill(C_SPACER)

    # Freeze rows 1–4
    ws.freeze_panes = "A5"

    # -------------------------------------------------------------------
    # Pillar sections (rows 5+)
    # -------------------------------------------------------------------
    input_cells = []        # populated as we write each question row
    dv_cells = {}           # validation_key -> cell address

    current_row = 5
    border = thin_border()

    for pillar in PILLARS:
        # Pillar header
        ws.row_dimensions[current_row].height = 22
        ws.merge_cells(f"A{current_row}:C{current_row}")
        c = ws[f"A{current_row}"]
        c.value = f"PILLAR {pillar['number']}  —  {pillar['title'].upper()}"
        c.fill = fill(C_PILLAR_BG)
        c.font = font_pillar()
        c.alignment = align_middle()
        current_row += 1

        for q in pillar["questions"]:
            ws.row_dimensions[current_row].height = 65

            # Col A — label + question
            ca = ws[f"A{current_row}"]
            ca.value = f"{q['label']}\n\n{q['prompt']}"
            ca.fill = fill(C_QUESTION_BG)
            ca.font = font_question()
            ca.alignment = align_wrap(indent=1)

            # Col B — input cell
            cb = ws[f"B{current_row}"]
            cb.value = q["placeholder"]
            cb.fill = fill(C_INPUT_BG)
            cb.font = font_input()
            cb.alignment = align_wrap()
            cb.border = border

            # Col C — guidance
            cc = ws[f"C{current_row}"]
            cc.value = q["guidance"]
            cc.fill = fill(C_GUIDANCE_BG)
            cc.font = font_guidance()
            cc.alignment = align_wrap(indent=1)

            # Track input cell address
            cell_addr = f"B{current_row}"
            input_cells.append(cell_addr)
            if q["validation"]:
                dv_cells[q["validation"]] = cell_addr

            current_row += 1

        # Spacer after each pillar
        ws.row_dimensions[current_row].height = 10
        for col in ["A", "B", "C"]:
            ws[f"{col}{current_row}"].fill = fill(C_SPACER)
        current_row += 1

    # -------------------------------------------------------------------
    # Patch completeness formula now that all row numbers are known
    # -------------------------------------------------------------------
    ws["A3"].value = build_completeness_formula(input_cells)

    # -------------------------------------------------------------------
    # Data validations
    # -------------------------------------------------------------------
    delivery_dv = DataValidation(
        type="list",
        formula1='"Done-for-you,Done-with-you,Self-serve,Hybrid,Mixed"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Selection",
        error="Please select a delivery model from the list provided.",
        showInputMessage=True,
        promptTitle="Delivery Model",
        prompt="Select the model that best describes your service delivery.",
    )
    delivery_dv.add(dv_cells["delivery_model"])
    ws.add_data_validation(delivery_dv)

    sector_dv = DataValidation(
        type="list",
        formula1='"Healthcare,Aesthetics,Financial Services,Legal,Education,Coaching/Consulting,E-commerce,SaaS/Tech,Other"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Selection",
        error="Please select your sector from the list provided.",
        showInputMessage=True,
        promptTitle="Sector",
        prompt="Select the sector that best describes your business.",
    )
    sector_dv.add(dv_cells["sector"])
    ws.add_data_validation(sector_dv)

    # -------------------------------------------------------------------
    # Save
    # -------------------------------------------------------------------
    out = os.path.normpath(OUTPUT_PATH)
    wb.save(out)
    print(f"Template written to: {out}")
    print(f"Input cells ({len(input_cells)}): {', '.join(input_cells)}")


if __name__ == "__main__":
    build_template()
